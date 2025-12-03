"""
Excel service for parsing batch data from XLSX files
"""
import os
from typing import List, Dict, Any, Optional
from pathlib import Path
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from ..errors import PDFFormFillerError


class ExcelError(PDFFormFillerError):
    """Excel-related errors"""
    pass


class ExcelService:
    """Service for parsing Excel files for batch processing"""

    @staticmethod
    def parse_batch_file(file_path: str) -> List[Dict[str, Any]]:
        """
        Parse Excel file and extract batch data

        Expected format:
        - First row: Header with field names
        - Subsequent rows: Data for each form to fill

        Args:
            file_path: Path to Excel file

        Returns:
            List of dictionaries, one per row

        Raises:
            ExcelError: If file cannot be parsed
        """
        if not os.path.exists(file_path):
            raise ExcelError(f"File not found: {file_path}")

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = workbook.active

            if sheet is None:
                raise ExcelError("No active sheet found in workbook")

            return ExcelService._parse_sheet(sheet)

        except openpyxl.utils.exceptions.InvalidFileException as e:
            raise ExcelError(f"Invalid Excel file: {e}")
        except Exception as e:
            raise ExcelError(f"Failed to parse Excel file: {e}")

    @staticmethod
    def _parse_sheet(sheet: Worksheet) -> List[Dict[str, Any]]:
        """
        Parse worksheet into list of dictionaries

        Args:
            sheet: Worksheet to parse

        Returns:
            List of dictionaries with form data
        """
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) < 2:
            raise ExcelError("Excel file must have at least 2 rows (header + data)")

        # First row is header
        headers = rows[0]

        if not headers or all(h is None for h in headers):
            raise ExcelError("First row must contain field names")

        # Clean headers (remove None values and empty strings)
        cleaned_headers = []
        for h in headers:
            if h is not None and str(h).strip():
                cleaned_headers.append(str(h).strip())
            else:
                cleaned_headers.append(None)

        # Parse data rows
        data_rows = []
        for row_idx, row in enumerate(rows[1:], start=2):
            row_data = {}
            has_data = False

            for col_idx, (header, value) in enumerate(zip(cleaned_headers, row)):
                # Skip columns without header
                if header is None:
                    continue

                # Convert value to appropriate type
                if value is not None:
                    # Handle boolean values
                    if isinstance(value, bool):
                        row_data[header] = value
                    # Handle numeric values
                    elif isinstance(value, (int, float)):
                        # Check if it's a whole number
                        if isinstance(value, float) and value.is_integer():
                            row_data[header] = str(int(value))
                        else:
                            row_data[header] = str(value)
                    # Handle string values
                    else:
                        str_value = str(value).strip()
                        if str_value:
                            # Try to detect boolean strings
                            lower_value = str_value.lower()
                            if lower_value in ('true', 'yes', 'sim', 'verdadeiro', '1', 'x'):
                                row_data[header] = True
                            elif lower_value in ('false', 'no', 'não', 'nao', 'falso', '0', ''):
                                row_data[header] = False
                            else:
                                row_data[header] = str_value
                            has_data = True

            # Only add row if it has at least one non-empty value
            if has_data:
                data_rows.append(row_data)

        if not data_rows:
            raise ExcelError("No data rows found in Excel file")

        return data_rows

    @staticmethod
    def create_template(field_names: List[str], output_path: str) -> str:
        """
        Create an Excel template with field names as headers

        Args:
            field_names: List of field names for the header
            output_path: Path to save the template

        Returns:
            Path to created template file

        Raises:
            ExcelError: If template cannot be created
        """
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Batch Data"

            # Write header row
            for col_idx, field_name in enumerate(field_names, start=1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.value = field_name
                # Make header bold
                cell.font = openpyxl.styles.Font(bold=True)

            # Add example row with placeholders
            for col_idx, field_name in enumerate(field_names, start=1):
                cell = sheet.cell(row=2, column=col_idx)
                cell.value = f"[{field_name}]"

            # Adjust column widths
            for col_idx in range(1, len(field_names) + 1):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

            # Save workbook
            workbook.save(output_path)

            return output_path

        except Exception as e:
            raise ExcelError(f"Failed to create template: {e}")

    @staticmethod
    def validate_batch_data(
        data: List[Dict[str, Any]],
        required_fields: Optional[List[str]] = None
    ) -> tuple[bool, List[str]]:
        """
        Validate batch data

        Args:
            data: List of dictionaries to validate
            required_fields: Optional list of required field names

        Returns:
            Tuple of (is_valid, list_of_errors)
        """
        errors = []

        if not data:
            errors.append("No data to validate")
            return False, errors

        # Check if required fields are present in all rows
        if required_fields:
            for row_idx, row in enumerate(data, start=1):
                missing_fields = []
                for field in required_fields:
                    if field not in row or row[field] is None or row[field] == '':
                        missing_fields.append(field)

                if missing_fields:
                    errors.append(
                        f"Row {row_idx}: Missing required fields: {', '.join(missing_fields)}"
                    )

        return len(errors) == 0, errors
